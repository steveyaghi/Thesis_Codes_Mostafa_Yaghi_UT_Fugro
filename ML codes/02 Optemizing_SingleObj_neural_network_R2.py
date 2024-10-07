import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_squared_error, r2_score
import tensorflow as tf
from sklearn.preprocessing import StandardScaler
from tensorflow import keras
from tensorflow.keras import layers
import tensorflow_probability as tfp
import seaborn as sns
from tensorflow.keras import regularizers
import optuna
from optuna.visualization import plot_optimization_history, plot_param_importances
from sklearn.metrics import r2_score, mean_squared_error
import inspect
from optuna.visualization import plot_contour, plot_edf, plot_parallel_coordinate 
import openpyxl
from openpyxl import load_workbook



#load the data



CC = pd.read_excel('clay_compaction_ML.xlsx')
CI = pd.read_excel('Clay_ins_ML.xlsx')
CIA = pd.read_excel('Autoencoded_CI.xlsx')
CT = pd.read_excel('Clay_Tri_F.xlsx')
COM = pd.read_excel('COMBINED_RF_dryunit_drydensity.xlsx')


#dropping unimportant columns for the performance of the model
CI = CI.drop(columns=['Unnamed: 0',
                      'clay_content',
                      'silt_content',
                      'sand_content',
                      'activity',
                      'activity_class',
                      'x',
                      'y',
                      'salt_content',
                      'mass_loss',
                      'void',
                      'organic_content',])





encoded_activity = pd.get_dummies(COM['activity_class'], prefix='activity')

# Step 2: Concatenate the resulting one-hot encoded DataFrame with the original DataFrame
COM = pd.concat([COM, encoded_activity], axis=1)

# Step 3 (Optional): Combine the one-hot encoded columns into a single column
# This will create a binary string representing the one-hot encoded values
COM['encoded_activity'] = encoded_activity.apply(lambda x: ''.join(x.astype(str)), axis=1)

def convert_to_binary_string(s):
    # Replace 'True' with '1' and 'False' with '0'
    binary_string = s.replace('True', '1').replace('False', '0')
    return binary_string

COM['encoded_activity'] = COM['encoded_activity'].apply(convert_to_binary_string)

COM = COM.drop(columns=['void',
                        'organic_content',
                        'activity_Active',
                        'activity_Highly Active',
                        'activity_Inactive',
                        'activity_Normal',
                        'activity_class'])


def convert_to_binary_string(s):
    # Replace 'True' with '1' and 'False' with '0'
    binary_string = s.replace('True', '1').replace('False', '0')
    return binary_string

# Apply the conversion function to each row in the 'class_1_encoded' column
COM['class_1_encoded'] = COM['class_1_encoded'].apply(convert_to_binary_string)





CC = CC[CC['degree_of_compaction'] >= 97]

CC = CC.drop(columns=['file_name',
                      'W_max',
                      'degree_of_compaction',
                      'mass'])



unique_classes_CI = CI['class%'].unique()

unique_classes_CT = CT['class 1'].unique()


CI = CI[CI['class%'] != 'other']

CT['q'] = CT['q'].fillna(CT['sigma_1'] - CT['sigma_3'])
# On-hot encoding for the clay classes


# One-hot encode the 'class 1' column
CI_encoded = pd.get_dummies(CI['class%'])

# Concatenate the original DataFrame with the one-hot encoded DataFrame
CI = pd.concat([CI, CI_encoded], axis=1)

# Convert one-hot columns to 0/1
for col in CI_encoded.columns:
    CI[col] = CI[col].astype(int)

# Combine one-hot columns into a single column
CI['class_1_encoded'] = CI[CI_encoded.columns].apply(lambda x: ''.join(x.astype(str)), axis=1)

# Drop the original one-hot encoded columns and the original class column
CI.drop(CI_encoded.columns, axis=1, inplace=True)
CI.drop('class%', axis=1, inplace=True)



#applying the encoding of Ins files to the triaxial data

# Define the mapping of class1 values to new column values
class1_mapping = {
    'ks1': '1000000',
    'ks2': '0100000',
    'ks3': '0010000',
    'ks4': '0001000',
    'kz1': '0000100',
    'kz2': '0000010',
    'kz3': '0000001'
}

# Create the new column using the map method
CT['class_encoded'] = CT['class 1'].map(class1_mapping)
CT.drop('class 1', axis=1, inplace=True)

# %%

def objective(trial, data_frame, dependent_var_columns, independent_var_columns):
    # Suggest hyperparameters
    learning_rate = trial.suggest_float('learning_rate', 1e-4, 1e-3, log=True)
    epochs = trial.suggest_int('epochs', 800, 1200)
    num_layers = trial.suggest_int('num_layers', 5, 12)
    units_per_layer = trial.suggest_int('units_per_layer', 150, 300)
    activation_functions = ['relu']
    activation = trial.suggest_categorical('activation', activation_functions)
    dropout_rate = trial.suggest_float('dropout_rate', 0.3, 0.55)
    batch_size = trial.suggest_int('batch_size', 32, 128)

    # Check and handle NaN values
    if data_frame.isnull().values.any():
        data_frame = data_frame.fillna(method='ffill').fillna(method='bfill') # Forward fill then backward fill
    # Covert Dataframe to numpy array
    data_array = data_frame.values
    #Extract dependent and independent variables using the column names
    independent_vars = data_frame[independent_var_columns].values
    dependent_vars = data_frame[dependent_var_columns].values
    #Identify encoded columns (replace with original values after scaling)
    encoded_columns = [i for i, col in enumerate(independent_var_columns) if 'encoded' in col]
    #exclude encoded columns from scaling
    independent_vars_to_scale = np.delete(independent_vars, encoded_columns, axis=1)
    dependent_vars_to_scale = dependent_vars
    #standardize the non-encoded columns
    scaler_independent = StandardScaler()
    scaler_dependent = StandardScaler()
    independent_vars_scaled = scaler_independent.fit_transform(independent_vars_to_scale)
    dependent_vars_scaled = scaler_dependent.fit_transform(dependent_vars_to_scale)

    for idx in encoded_columns:
        independent_vars_scaled = np.insert(independent_vars_scaled, idx, independent_vars[:, idx], axis=1)

    X_train, X_test, y_train, y_test = train_test_split(independent_vars_scaled, dependent_vars_scaled, test_size=0.2, random_state=42)
    #build the nwural network model
    model = tf.keras.Sequential() 
    model.add(tf.keras.Input(shape=(X_train.shape[1],)))
    for _ in range(num_layers - 1):
        model.add(layers.Dense(units_per_layer, activation=activation,
                               kernel_regularizer=regularizers.l2(1e-4)))
        model.add(layers.BatchNormalization())  # Add Batch Normalization
    model.add(layers.Dense(len(dependent_var_columns)))
    
    #compile the model
    model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=learning_rate), loss='mean_squared_error')

    # Add ReduceLROnPlateau and EarlyStopping callbacks
    reduce_lr = tf.keras.callbacks.ReduceLROnPlateau(monitor='val_loss', factor=0.5, patience=10, min_lr=1e-6)
    early_stopping = tf.keras.callbacks.EarlyStopping(monitor='val_loss', patience=50, restore_best_weights=True)

    #train the model
    history = model.fit(X_train, y_train, epochs=epochs, validation_split=0.2, verbose=0)

    training_loss = history.history['loss']
    validation_loss = history.history['val_loss']

    y_pred_scaled = model.predict(X_test)
    if np.isnan(y_pred_scaled).any():
        raise ValueError("Model predictions contain NaN values.")
    
    y_pred = scaler_dependent.inverse_transform(y_pred_scaled)
    y_test_original = scaler_dependent.inverse_transform(y_test)

    r2_scores = [r2_score(y_test_original[:, i], y_pred[:, i]) for i in range(y_test_original.shape[1])]
    mean_r2 = np.mean(r2_scores)

    print("Individual R2 scores:", r2_scores) # print Individual R2 scores
    
    return mean_r2, training_loss, validation_loss, r2_scores

def optimize_neural_network(data_frame, dependent_var_columns, independent_var_columns, n_trials=50, excel_file_path='Optuna_scores.xlsx'):
    for var_name, var_value in inspect.currentframe().f_back.f_locals.items():
        if var_value is data_frame:
            data_name = var_name
            break

    study_name = f"CC Single objective 8 HP, +R2, {n_trials}-trials_{data_name}-USING_{','.join(independent_var_columns)}"
    storage = optuna.storages.RDBStorage('sqlite:///db.sqlite3')

    try:
        study = optuna.load_study(study_name=study_name, storage=storage)
    except KeyError:
        study = optuna.create_study(study_name=study_name, direction='maximize', storage=storage)
    else:
        study_name = "AUTOADD_" + study_name
        study = optuna.create_study(study_name=study_name, direction='maximize', storage=storage)

    training_losses = []
    validation_losses = []
    all_r2_scores = []

    def objective_with_logging(trial):
        result, training_loss, validation_loss, r2_scores = objective(trial, data_frame, dependent_var_columns, independent_var_columns)
        
        # Log individual R² scores for each predicted variable
        print(f"Trial {trial.number} R² scores for each variable:")
        for var_name, r2_score_value in zip(dependent_var_columns, r2_scores):
            print(f"{var_name}: {r2_score_value:.4f}")




        
        training_losses.append(training_loss)
        validation_losses.append(validation_loss)
        all_r2_scores.append(r2_scores)  # collect r2 scores
        append_to_excel(excel_file_path, study_name, r2_scores, dependent_var_columns, independent_var_columns)
        return result

    study.optimize(objective_with_logging, n_trials=n_trials)

    plot_optimization_history(study)
    plot_param_importances(study)

    best_trial = study.best_trial

    print("Best hyperparameters:", study.best_params)
    return study, training_losses, validation_losses, all_r2_scores

def plot_loss_histories(training_losses, validation_losses):
    plt.figure(figsize=(12, 6))
    
    # Plot the losses of the best trial
    plt.plot(training_losses[best_trial.number], label='Best Trial Training Loss')
    plt.plot(validation_losses[best_trial.number], label='Best Trial Validation Loss', linestyle='dashed')
    
    plt.xlabel('Epoch')
    plt.ylabel('Loss')
    plt.title('Training and Validation Loss + study_name')
    plt.legend(loc='upper right')
    plt.show()

def visualize_optuna_results(study, training_losses, validation_losses):
    optimization_history_fig = plot_optimization_history(study)
    param_importances_fig = plot_param_importances(study)
    contour_fig = plot_contour(study, params=["learning_rate", "epochs"])
    edf_fig = plot_edf(study)
    parallel_coordinate_fig = plot_parallel_coordinate(study, params=["learning_rate", "epochs", "num_layers"])

    optuna.visualization.plot_optimization_history(study)
    optuna.visualization.plot_param_importances(study)
    optuna.visualization.plot_contour(study, params=["learning_rate", "epochs"])
    optuna.visualization.plot_edf(study)
    optuna.visualization.plot_parallel_coordinate(study, params=["learning_rate", "epochs", "num_layers"])
    
    plot_loss_histories(training_losses, validation_losses)

def plot_individual_r2_scores(r2_scores, dependent_var_columns):
    mean_r2_scores = np.mean(r2_scores, axis=0)
    plt.bar(dependent_var_columns, mean_r2_scores)
    plt.xlabel('Dependent Variable')
    plt.ylabel('Mean R2 Score')
    plt.title('Mean R2 Score for each Dependent Variable')
    plt.show()

def append_to_excel(file_path, study_name, r2_scores, dependent_var_columns, independent_var_columns):
    try:
        workbook = load_workbook(file_path)
        if study_name in workbook.sheetnames:
            sheet = workbook[study_name]
        else:
            sheet = workbook.create_sheet(study_name)
            # Add header row if the sheet is newly created
            sheet.append(['Study Name'] + dependent_var_columns + independent_var_columns)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = study_name
        # Add header row if the workbook is newly created
        sheet.append(['Study Name'] + dependent_var_columns + independent_var_columns)

    # Add the new row of data
    row = [study_name] + r2_scores + independent_var_columns
    sheet.append(row)
    workbook.save(file_path)
#%%
# Update the function call
#independent_var_columns = ['clay_content_100%', 'sand_content_100%', 'silt_content_100%']
independent_var_columns = ['water_content']
#independent_var_columns = ['liquid_limit', 'plastic_limit', 'plasticity_index']
dependent_var_columns = [col for col in CC.columns if col not in independent_var_columns]

study, training_losses, validation_losses, all_r2_scores = optimize_neural_network(CC, dependent_var_columns,
                                                                     independent_var_columns,
                                                                       n_trials=100,
                                                                       excel_file_path='Optuna_scores.xlsx')
visualize_optuna_results(study, training_losses, validation_losses)
plot_individual_r2_scores(all_r2_scores, dependent_var_columns)
# %%

#independent_var_columns = ['dry_unit_weight']
#dependent_var_columns = [col for col in CT.columns if col not in independent_var_columns]

#study, training_losses, validation_losses, all_r2_scores = optimize_neural_network(CT, dependent_var_columns,
#                                                                     independent_var_columns,
#                                                                       n_trials=100,
#                                                                       excel_file_path='Optuna_scores.xlsx')
#visualize_optuna_results(study, training_losses, validation_losses)
#plot_individual_r2_scores(all_r2_scores, dependent_var_columns)